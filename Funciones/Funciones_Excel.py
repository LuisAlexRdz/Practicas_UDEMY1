import openpyxl

class Funexcel():
    def __init__(self, driver):
        self.driver = driver

#Funcion para contar las filas que contiene el documento excel
    def getRowCount(file,path,sheetname):
        Workbook = openpyxl.load_workbook(path) #manda llamar la libreria y carga del libro la ruta
        sheet = Workbook[sheetname] #Manda llamar la hoja del libro
        return (sheet.max_row) #cuenta el maximo de filas

#Funcion para contar las columnas que contiene el documento excel
    def getColumnCount(file,sheetname):
        Workbook = openpyxl.load_workbook(file) #manda llamar la libreria y carga del libro la ruta
        sheet = Workbook[sheetname] #Manda llamar la hoja del libro
        return (sheet.max_column) #cuenta el maximo de columnas

#Funcion para leer los datos almacenados en la  hoja de excel
    def readData(file, path, sheetname, rownum, columnnum):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetname]
        return sheet.cell(row=rownum, column=columnnum).value #Guarda en una variable el valor del renglon y guarda en una variable el valor de la columna

#Funcion para excribir en la hoja del documento excel
    def writeData(file, path, sheetname, rownum, columnnum, data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetname]
        sheet.cell(row=rownum, column=columnnum).value = data #Guarda en una variable el valor del renglon y guarda en una variable el valor de la columna
        Workbook.save(path)